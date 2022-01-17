import cx_Oracle
from cx_Oracle import Cursor
from openpyxl import load_workbook

cx_Oracle.init_oracle_client(lib_dir=r"C:\SN5\Oracle\instantclient_21_3")

class Dao:

    def __init__(self):
        self.connection = cx_Oracle.connect(

            user="click", password="falaae",
            dsn="kanto/orcl",
            encoding="UTF-8"
        )

    def existeNumero(self, numero):
        cursor: Cursor = self.connection.cursor()

        lista = list(cursor.execute(f"SELECT NUMERO FROM ENVIO_WHATSAPP WHERE NUMERO = {numero}"))

        if len(lista) > 0:
            return True

        return False

    def criaNumero(self, cod_campanha, numero):
        cursor: Cursor = self.connection.cursor()

        cursor.execute(
            f"INSERT INTO \"CLICK\".\"ENVIO_WHATSAPP\" (COD_CAMPANHA, NUMERO) VALUES ('{cod_campanha}', '{numero}')")
        self.connection.commit()


# cod_campanha = 1
# xlsx_path = r"C:\Users\DEV\Desktop\CalceContatos\RESPOSTA GUARA.xlsx"
#
# xlsx = load_workbook(xlsx_path)
# sheet = xlsx.active
# max_row = sheet.max_row
#
# dao = Dao()
#
# print(dao.existeNumero(61981351430))
#
# for r in range(1, max_row + 1):
#
#     cell_num = sheet.cell(column=2, row=r)
#     cell_ok = sheet.cell(column=3, row=r)
#
#     print(cell_num.value)
#
#     if cell_ok.value == "Sim":
#
#         if not dao.existeNumero(cell_num.value):
#             dao.criaNumero(cod_campanha, cell_num.value)
