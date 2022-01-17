import random
from time import sleep

from openpyxl import load_workbook

import WhatsApp
from WhatsApp.dao import Dao

texto = """
Olá {nome}

*CHEEGGGOOUU O NATAL!*
*CALCE PERFEITO*

Venha conhecer a 
a *Nova coleção* 
Usaflex, Piccadilly 
Opananken e Skechers.

Calçados leves, 
coloridos e festivos
para dar um 
*up no visual.*

E tem mais! Em compras 
a partir de R$ 400,00, 
*você GANHA*
 uma *linda toalha de banho* 
para arrasar no verão!* 
🎁 🛍️

Aguardamos você com 
café bem quentinho🥰

*Para mais informações, consulte o regulamento em alguma de nossas lojas."""

cod_campanha = 1
xlsx_path = r"C:\Users\DEV\Desktop\CalceContatos\Inauguração Calce Perfeito Guara.xlsx"
file_path = r""

xlsx = load_workbook(xlsx_path)
sheet = xlsx.active
max_row = sheet.max_row

dao = Dao()

for r in range(2, max_row + 1):

    cell_nome = sheet.cell(column=1, row=r)
    cell_num = sheet.cell(column=2, row=r)
    cell_ok = sheet.cell(column=3, row=r)

    print(f"{r} - {max_row}: {cell_num.value}", end=" ")

    if cell_nome.value is not None:
        print(f"{cell_nome.value}", end=" ")

    if cell_ok.value is not None:
        print(f"{cell_ok.value}", end=" ")

    numero = str(cell_num.value).replace("(", "")

    if len(numero) == 9:
        numero = "61" + numero

    if len(numero) == 10:
        numero = "619" + numero[2:]

    if cell_ok.value != "Sim" and cell_ok.value != "Nao":

        if not dao.existeNumero(numero):

            #Retirando 619
            numero = str(numero)[3:]

            if WhatsApp.pesquisar(numero):

                sleep(0.2)

                nome = str(cell_nome.value)
                nome = nome.split()[0].capitalize()

                tx = texto.replace("{nome}", nome)

                WhatsApp.enviarTX(tx)

                cell_ok.value = "Sim"
                print("Sim", end=" ")
                xlsx.save(xlsx_path)

                dao.criaNumero(cod_campanha, cell_num.value)

                sleep(random.randrange(30, 40))

            else:

                cell_ok.value = "Não"
                print("Não", end=" ")
                xlsx.save(xlsx_path)

    print("")

print("FIM")
