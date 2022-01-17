from openpyxl import load_workbook

from WhatsAppTx import WhatsApp

path_exel = r"C:\Users\DEV\Desktop\CalceContatos\RESPOSTA NORTE.xlsx"
arquivo_excel = load_workbook(path_exel)
plan = arquivo_excel.active

# Criando cabeçario
if plan.cell(column=5, row=1).value != "RESPOSTA":

    plan.insert_rows(1)
    plan.cell(column=1, row=1).value = "NOME"
    plan.cell(column=2, row=1).value = "NUMERO"
    plan.cell(column=3, row=1).value = "CONTATO ENCOTRADO"
    plan.cell(column=4, row=1).value = "MENSAGEM ENTREGUE"
    plan.cell(column=5, row=1).value = "RESPOSTA"

    plan.column_dimensions['A'].width = 50
    plan.column_dimensions['B'].width = 20
    plan.column_dimensions['C'].width = 22
    plan.column_dimensions['D'].width = 22
    plan.column_dimensions['E'].width = 50

    arquivo_excel.save(path_exel)

max_row = plan.max_row

TEXTO = "CHEEGGGOOUU"

ws = WhatsApp()

for r in range(1, max_row + 1):

    cell_num = plan.cell(column=2, row=r)
    cell_ok = plan.cell(column=3, row=r)
    cell_entregue = plan.cell(column=4, row=r)
    cell_res = plan.cell(column=5, row=r)

    if cell_ok.value == "Sim":

        if cell_entregue.value == None:

            numero = str(cell_num.value).replace("+", "").replace(" ", "").replace("-", "").replace("(", "").replace(")", "")

            if len(numero) == 11:
                numero = "55" + numero

            print(f"{numero}", end=" ")

            if ws.getContact(numero):

                r = ws.wasAnswered(TEXTO)

                cell_entregue.value = r[0]
                cell_res.value = r[1]
            else:
                print("Contato não encontrado", end=" ")
                cell_entregue.value = "Nao"
                cell_res.value = ""

            print(f"= E: {cell_entregue.value} R: {cell_res.value}")
            arquivo_excel.save(path_exel)
